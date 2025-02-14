import cv2
import numpy as np
import os
from collections import defaultdict
import openpyxl

def process_video(video_source):
    video_name = os.path.splitext(os.path.basename(video_source))[0]
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "MarkerIDs"

    up_col = 1
    down_col = 1

    os.makedirs(video_name, exist_ok=True)
    excel_filename = f"{video_name}.xlsx"

    aruco_dict = cv2.aruco.getPredefinedDictionary(cv2.aruco.DICT_4X4_1000)
    parameters = cv2.aruco.DetectorParameters()
    detector = cv2.aruco.ArucoDetector(aruco_dict, parameters)

    cap = cv2.VideoCapture(video_source)
    if not cap.isOpened():
        print(f"Error: Cannot open video {video_source}")
        return

    consecutive_counts = defaultdict(int)
    saved_ids_up = set()
    saved_ids_down = set()

    up_save_index = 1
    down_save_index = 1

    while True:
        ret, frame = cap.read()
        if not ret:
            break 

        height, width = frame.shape[:2]
        mid_x = width // 2  

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        corners, ids, _ = detector.detectMarkers(gray)

        if ids is not None:
            cv2.aruco.drawDetectedMarkers(frame, corners, ids)

            for marker_id_arr, corner in zip(ids, corners):
                marker_id = marker_id_arr[0]

                cx = np.mean(corner[0][:, 0])
                zone_label = "UP" if cx < mid_x else "DOWN"

                if zone_label == "UP" and marker_id in saved_ids_up:
                    continue
                if zone_label == "DOWN" and marker_id in saved_ids_down:
                    continue

                consecutive_counts[(marker_id, zone_label)] += 1

                if consecutive_counts[(marker_id, zone_label)] == 3:
                    if zone_label == "UP":
                        file_name = f"UP_{up_save_index}_{marker_id}.png"
                        up_save_index += 1
                    else:
                        file_name = f"DOWN_{down_save_index}_{marker_id}.png"
                        down_save_index += 1

                    save_path = os.path.join(video_name, file_name)
                    cv2.imwrite(save_path, frame)
                    print(f"Saved: {save_path}")

                    if zone_label == "UP":
                        saved_ids_up.add(marker_id)
                    else:
                        saved_ids_down.add(marker_id)

                    if zone_label == "UP":
                        sheet.cell(row=1, column=up_col).value = marker_id
                        up_col += 1
                    else:
                        sheet.cell(row=2, column=down_col).value = marker_id
                        down_col += 1

                    wb.save(excel_filename)

                    consecutive_counts[(marker_id, zone_label)] = 0

        else:
            for key in list(consecutive_counts.keys()):
                consecutive_counts[key] = 0

    # cap.release()
    # cv2.destroyAllWindows()
    print(f"Processing finished for {video_source}")

    wb.save(excel_filename)
    print(f"Excel saved at {excel_filename}")


def main(video_sources):

    for vs in video_sources:
        process_video(vs)


if __name__ == "__main__":
    videos = [
        r"C:\Users\phet\Downloads\VID_20250213_145926.mp4",
        r"C:\Users\phet\Downloads\VID_20250213_150110.mp4"
    ]
    main(videos)